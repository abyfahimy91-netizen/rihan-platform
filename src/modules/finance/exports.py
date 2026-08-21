"""
Export Services برای ماژول مالی (US-031)

خروجی اکسل از گزارش‌های مالی
"""
from io import BytesIO
from decimal import Decimal
import jdatetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .models import SupplierLedger, SupplierTransaction, Settlement
from .services import FinanceService


class FinanceExporter:
    """سرویس export گزارش‌های مالی به Excel"""

    @staticmethod
    def export_all_transactions():
        """Export تمام تراکنش‌ها به اکسل"""
        wb = Workbook()
        ws = wb.active
        ws.title = 'تراکنش‌های مالی'
        ws.sheet_view.rightToLeft = True  # RTL

        # Header styling
        header_fill = PatternFill(start_color='2D5A2D', end_color='2D5A2D', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        header_align = Alignment(horizontal='center', vertical='center')

        # Headers
        headers = ['تاریخ (شمسی)', 'تأمین‌کننده', 'نوع تراکنش', 'مبلغ (تومان)', 
                   'شماره سفارش', 'توضیحات', 'ثبت‌کننده']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        # Data
        transactions = SupplierTransaction.objects.select_related(
            'ledger__supplier', 'order', 'created_by'
        ).order_by('-created_at')

        for row, tx in enumerate(transactions, 2):
            # تبدیل تاریخ میلادی به شمسی
            jalali_date = jdatetime.date.fromgregorian(date=tx.created_at.date())
            jalali_str = jalali_date.strftime('%Y/%m/%d')

            ws.cell(row=row, column=1, value=jalali_str)
            ws.cell(row=row, column=2, value=tx.ledger.supplier.title)
            ws.cell(row=row, column=3, value=tx.get_transaction_type_display())
            ws.cell(row=row, column=4, value=float(tx.amount))
            ws.cell(row=row, column=5, value=tx.order.order_number if tx.order else '-')
            ws.cell(row=row, column=6, value=tx.description or '-')
            ws.cell(row=row, column=7, value=tx.created_by.username if tx.created_by else 'سیستم')

        # Auto-width columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        # Format amount column as number
        for row in range(2, len(transactions) + 2):
            ws.cell(row=row, column=4).number_format = '#,##0'

        # Summary sheet
        ws2 = wb.create_sheet('خلاصه')
        ws2.sheet_view.rightToLeft = True

        ws2['A1'] = 'خلاصه گزارش مالی'
        ws2['A1'].font = Font(bold=True, size=14, color='2D5A2D')

        stats = FinanceService.get_dashboard_stats(days=30)
        ws2['A3'] = 'درآمد ۳۰ روز گذشته:'
        ws2['B3'] = float(stats['total_revenue'])
        ws2['A4'] = 'تعداد سفارش‌های تحویل شده:'
        ws2['B4'] = stats['order_count']
        ws2['A5'] = 'میانگین ارزش سفارش:'
        ws2['B5'] = float(stats['avg_order_value'])
        ws2['A6'] = 'مجموع بدهی به تأمین‌کنندگان:'
        ws2['B6'] = float(stats['total_supplier_debt'])

        for row in [3, 4, 5, 6]:
            ws2.cell(row=row, column=1).font = Font(bold=True)
            if row in [3, 5, 6]:
                ws2.cell(row=row, column=2).number_format = '#,##0'

        ws2.column_dimensions['A'].width = 35
        ws2.column_dimensions['B'].width = 20

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
